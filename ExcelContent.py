from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile
import logging
from Config import *

current_date = datetime.now().strftime('%d/%m/%Y')

def create_excel(rows_query1, rows_query2):
    dicts_query1 = [{'LeadId': row[0], 'INTERNET_ENQUIRY_ID': row[1], 'ems_server_response': row[2], 'name': row[3], 'create_date_get': row[4]} for row in rows_query1]

    if not rows_query2:
        dicts_query2 = [{'LeadId': '', 'INTERNET_ENQUIRY_ID': '', 'ems_server_response': '', 'name': '','dms_response':'', 'create_date': ''}]
        
    else:
        dicts_query2 = [{'LeadId': row[0], 'INTERNET_ENQUIRY_ID': row[1], 'ems_server_response': row[2], 'name': row[3],'dms_response': row[4], 'create_date': row[5]} for row in rows_query2]

    #local_path = "D:TVS_LMS_Test.xlsx"

    #_, excel_file_name = local_path, f"TVS_LMS_{current_date.replace('/', '_')}.xlsx"
    _, excel_file_name = tempfile.mkstemp(suffix=".xlsx", prefix=f"TVS_LMS_{current_date.replace('/', '_')}")

    wb = Workbook()

    ws_combined = wb.active
    ws_combined.title = "Ems Repush(91Wheels)"
    sky_blue_fill = PatternFill(start_color='7EC0EE', end_color='7EC0EE', fill_type='solid')
    ws_combined.merge_cells('A1:E1')
    ws_combined['A1'] = "EMS_Success_response"
    ws_combined.merge_cells('G1:L1')
    ws_combined['G1'] = "CRM_Success_response"
    header_combined = ["LeadId", "INTERNET_ENQUIRY_ID", "ems_server_response", "name", "create_date_get", "", "LeadId", "INTERNET_ENQUIRY_ID", "ems_server_response", "name", "dms_response","create_date"]
    ws_combined.append(header_combined)

    m1 = ws_combined['A2:E2'][0]
    for cell in m1:
        cell.fill = sky_blue_fill
    m2 = ws_combined['G2:L2'][0]
    for cell in m2:
        cell.fill = sky_blue_fill

    for idx, row1 in enumerate(dicts_query1, start=2):  # Start from row 2 in column G
        if not dicts_query2 or idx - 2 >= len(dicts_query2):
            combined_row = [row1[key] if key != '' else '' for key in header_combined[:5] + [''] * 6]
        else:
            row2 = dicts_query2[idx - 2]
            combined_row1 = [row2[key] if key != '' else '' for key in header_combined[6:]]
            combined_row = [row1[key] if key != '' else '' for key in header_combined[:5]]
            combined_row.extend([''] )  # Add empty column
            combined_row.extend(combined_row1)
        ws_combined.append(combined_row)


    for column in ws_combined.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_combined.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    border_style = openpyxl.styles.Side(style='thin')
    border = openpyxl.styles.Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    for row in ws_combined.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(excel_file_name)

    print("Excel file saved:", excel_file_name)
    logging.info(f"Excel file saved: {excel_file_name}")
    return excel_file_name
